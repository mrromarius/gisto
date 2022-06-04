import sys
from PyQt5.QtWidgets import QApplication,  QMessageBox
from PyQt5 import QtWidgets
from PyQt5.QtCore import QRect, Qt
from openpyxl import Workbook

from MainWindow import Ui_MainWindow
from Histogram import Ui_Histogram
from Ontology import Ui_Ontology
from QueryTable import Ui_QueryTable


class GistoWindow(QtWidgets.QMainWindow, Ui_Histogram):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.click_close)
        self.pushButton_gisto.clicked.connect(self.click_load_gisto)

    def click_close(self):

        self.close()

    def click_load_gisto(self):
        # хороший способ реализации, но тогда нужно полностью сделать генерацию формы
        # https://www.cyberforum.ru/python-graphics/thread1513449.html#post7966919
        
        # тут делаем жуткий хард-код потому что временя н еждет, можно потом переделать всю форму на динамическое создание
        ya_data = [int(self.lineEdit_1.text()),int(self.lineEdit_2.text()),int(self.lineEdit_3.text()),int(self.lineEdit_4.text()),int(self.lineEdit_5.text()), 
        int(self.lineEdit_6.text()),int(self.lineEdit_7.text()),int(self.lineEdit_8.text()),int(self.lineEdit_9.text()), int(self.lineEdit_10.text()),
        int(self.lineEdit_11.text()),int(self.lineEdit_12.text()),int(self.lineEdit_13.text()),int(self.lineEdit_14.text()),int(self.lineEdit_15.text()), 
        int(self.lineEdit_16.text()),int(self.lineEdit_17.text()),int(self.lineEdit_18.text()),int(self.lineEdit_19.text()), int(self.lineEdit_20.text()),
        int(self.lineEdit_21.text()),int(self.lineEdit_22.text()),int(self.lineEdit_23.text()),int(self.lineEdit_24.text()), int(self.lineEdit_25.text()),
        0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        gg_data = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,
        int(self.lineEdit_26.text()),int(self.lineEdit_27.text()),int(self.lineEdit_28.text()),int(self.lineEdit_29.text()), int(self.lineEdit_30.text()), 
        int(self.lineEdit_31.text()),int(self.lineEdit_32.text()),int(self.lineEdit_33.text()),int(self.lineEdit_34.text()), int(self.lineEdit_35.text()), 
        int(self.lineEdit_36.text()),int(self.lineEdit_37.text()),int(self.lineEdit_38.text()),int(self.lineEdit_39.text()), int(self.lineEdit_40.text()), 
        int(self.lineEdit_41.text()),int(self.lineEdit_42.text()),int(self.lineEdit_43.text()),int(self.lineEdit_44.text()), int(self.lineEdit_45.text()),
        int(self.lineEdit_46.text()),int(self.lineEdit_47.text()),int(self.lineEdit_48.text()),int(self.lineEdit_49.text()), int(self.lineEdit_50.text())]
        y = [i for i in range(1, 52)]
        self.graphWidget.clear()
        
        self.graphWidget.setGeometry(QRect(410, 70, 800, 600))

        self.graphWidget.plot(y,ya_data, name="Yandex", stepMode=True, fillLevel = 0, brush=(140,230,32,150))
        self.graphWidget.plot(y,gg_data, name="Google", stepMode=True, fillLevel = 0, brush=(255,219,139,180))
        self.graphWidget.addLegend()
        grid = QtWidgets.QGridLayout()
        grid.addWidget(self.graphWidget, 3, 4, Qt.AlignCenter)

class OntoWindow(QtWidgets.QMainWindow, Ui_Ontology):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.pushButton.clicked.connect(self.click_close)

    def click_close(self):
        self.close()

class TableWindow(QtWidgets.QMainWindow, Ui_QueryTable):
    def __init__(self, tableName):
        super().__init__()
        self.setupUi(self)
        self.tableName = tableName
        self.pushButton.clicked.connect(self.click_close)
        self.pushButton_3.clicked.connect(self.click_save_to_excel)

    def click_close(self):
        self.close()

    def click_save_to_excel(self):
        excel_file = Workbook()
        excel_sheet = excel_file.create_sheet(title = self.tableName, index = 0)
        for count in range(1,30, 7):
            excel_sheet.cell(row=count, column=1).value = "Запрос"
            excel_sheet.cell(row=count+1, column=1).value = "Источник"
            excel_sheet.cell(row=count+1, column=2).value = "Оценка"
            excel_sheet.cell(row=count+1, column=3).value = "Тип ресурса"

        excel_sheet['B1'] = self.textEditZapros1.toPlainText()
        excel_sheet['B8'] = self.textEditZapros2.toPlainText()
        excel_sheet['B15'] = self.textEditZapros3.toPlainText()
        excel_sheet['B22'] = self.textEditZapros4.toPlainText()
        excel_sheet['B29'] = self.textEditZapros5.toPlainText()

        excel_sheet.cell(row=3, column=1).value = self.textEdit_source_1.toPlainText()
        excel_sheet.cell(row=4, column=1).value = self.textEdit_source_2.toPlainText()
        excel_sheet.cell(row=5, column=1).value = self.textEdit_source_3.toPlainText()
        excel_sheet.cell(row=6, column=1).value = self.textEdit_source_4.toPlainText()
        excel_sheet.cell(row=7, column=1).value = self.textEdit_source_5.toPlainText()
        excel_sheet.cell(row=10, column=1).value = self.textEdit_source_6.toPlainText()
        excel_sheet.cell(row=11, column=1).value = self.textEdit_source_7.toPlainText()
        excel_sheet.cell(row=12, column=1).value = self.textEdit_source_8.toPlainText()
        excel_sheet.cell(row=13, column=1).value = self.textEdit_source_9.toPlainText()
        excel_sheet.cell(row=14, column=1).value = self.textEdit_source_10.toPlainText()
        excel_sheet.cell(row=17, column=1).value = self.textEdit_source_11.toPlainText()
        excel_sheet.cell(row=18, column=1).value = self.textEdit_source_12.toPlainText()
        excel_sheet.cell(row=19, column=1).value = self.textEdit_source_13.toPlainText()
        excel_sheet.cell(row=20, column=1).value = self.textEdit_source_14.toPlainText()
        excel_sheet.cell(row=21, column=1).value = self.textEdit_source_15.toPlainText()
        excel_sheet.cell(row=24, column=1).value = self.textEdit_source_16.toPlainText()
        excel_sheet.cell(row=25, column=1).value = self.textEdit_source_17.toPlainText()
        excel_sheet.cell(row=26, column=1).value = self.textEdit_source_18.toPlainText()
        excel_sheet.cell(row=27, column=1).value = self.textEdit_source_19.toPlainText()
        excel_sheet.cell(row=28, column=1).value = self.textEdit_source_20.toPlainText()
        excel_sheet.cell(row=31, column=1).value = self.textEdit_source_21.toPlainText()
        excel_sheet.cell(row=32, column=1).value = self.textEdit_source_22.toPlainText()
        excel_sheet.cell(row=33, column=1).value = self.textEdit_source_23.toPlainText()
        excel_sheet.cell(row=34, column=1).value = self.textEdit_source_24.toPlainText()
        excel_sheet.cell(row=35, column=1).value = self.textEdit_source_25.toPlainText()

        excel_sheet.cell(row=3, column=2).value = self.spinBox1.value()
        excel_sheet.cell(row=4, column=2).value = self.spinBox_2.value()
        excel_sheet.cell(row=5, column=2).value = self.spinBox_3.value()
        excel_sheet.cell(row=6, column=2).value = self.spinBox_4.value()
        excel_sheet.cell(row=7, column=2).value = self.spinBox_5.value()
        excel_sheet.cell(row=10, column=2).value = self.spinBox_6.value()
        excel_sheet.cell(row=11, column=2).value = self.spinBox_7.value()
        excel_sheet.cell(row=12, column=2).value = self.spinBox_8.value()
        excel_sheet.cell(row=13, column=2).value = self.spinBox_9.value()
        excel_sheet.cell(row=14, column=2).value = self.spinBox_10.value()
        excel_sheet.cell(row=17, column=2).value = self.spinBox_11.value()
        excel_sheet.cell(row=18, column=2).value = self.spinBox_12.value()
        excel_sheet.cell(row=19, column=2).value = self.spinBox_13.value()
        excel_sheet.cell(row=20, column=2).value = self.spinBox_14.value()
        excel_sheet.cell(row=21, column=2).value = self.spinBox_15.value()
        excel_sheet.cell(row=24, column=2).value = self.spinBox_16.value()
        excel_sheet.cell(row=25, column=2).value = self.spinBox_17.value()
        excel_sheet.cell(row=26, column=2).value = self.spinBox_18.value()
        excel_sheet.cell(row=27, column=2).value = self.spinBox_19.value()
        excel_sheet.cell(row=28, column=2).value = self.spinBox_20.value()
        excel_sheet.cell(row=31, column=2).value = self.spinBox_21.value()
        excel_sheet.cell(row=32, column=2).value = self.spinBox_22.value()
        excel_sheet.cell(row=33, column=2).value = self.spinBox_23.value()
        excel_sheet.cell(row=34, column=2).value = self.spinBox_24.value()
        excel_sheet.cell(row=35, column=2).value = self.spinBox_25.value()

        excel_sheet.cell(row=3, column=3).value = self.textEdit_type_1.toPlainText()
        excel_sheet.cell(row=4, column=3).value = self.textEdit_type_2.toPlainText()
        excel_sheet.cell(row=5, column=3).value = self.textEdit_type_3.toPlainText()
        excel_sheet.cell(row=6, column=3).value = self.textEdit_type_4.toPlainText()
        excel_sheet.cell(row=7, column=3).value = self.textEdit_type_5.toPlainText()
        excel_sheet.cell(row=10, column=3).value = self.textEdit_type_6.toPlainText()
        excel_sheet.cell(row=11, column=3).value = self.textEdit_type_7.toPlainText()
        excel_sheet.cell(row=12, column=3).value = self.textEdit_type_8.toPlainText()
        excel_sheet.cell(row=13, column=3).value = self.textEdit_type_9.toPlainText()
        excel_sheet.cell(row=14, column=3).value = self.textEdit_type_10.toPlainText()
        excel_sheet.cell(row=17, column=3).value = self.textEdit_type_11.toPlainText()
        excel_sheet.cell(row=18, column=3).value = self.textEdit_type_12.toPlainText()
        excel_sheet.cell(row=19, column=3).value = self.textEdit_type_13.toPlainText()
        excel_sheet.cell(row=20, column=3).value = self.textEdit_type_14.toPlainText()
        excel_sheet.cell(row=21, column=3).value = self.textEdit_type_15.toPlainText()
        excel_sheet.cell(row=24, column=3).value = self.textEdit_type_16.toPlainText()
        excel_sheet.cell(row=25, column=3).value = self.textEdit_type_17.toPlainText()
        excel_sheet.cell(row=26, column=3).value = self.textEdit_type_18.toPlainText()
        excel_sheet.cell(row=27, column=3).value = self.textEdit_type_19.toPlainText()
        excel_sheet.cell(row=28, column=3).value = self.textEdit_type_20.toPlainText()
        excel_sheet.cell(row=31, column=3).value = self.textEdit_type_21.toPlainText()
        excel_sheet.cell(row=32, column=3).value = self.textEdit_type_22.toPlainText()
        excel_sheet.cell(row=33, column=3).value = self.textEdit_type_23.toPlainText()
        excel_sheet.cell(row=34, column=3).value = self.textEdit_type_24.toPlainText()
        excel_sheet.cell(row=35, column=3).value = self.textEdit_type_25.toPlainText()

        excel_file.save(filename=self.tableName + ".xlsx")
        msg = QMessageBox()
        msg.setText("Данные сохранены в файле " + self.tableName + ".xlsx")
        msg.setWindowTitle("Information")
        msg.exec_()

class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

        self.ui.pushButton_5.clicked.connect(self.click_ontol)
        self.ui.pushButton_6.clicked.connect(self.click_gisto)
        self.ui.pushButton.clicked.connect(self.click_table1)
        self.ui.pushButton_2.clicked.connect(self.click_table2)
        self.ui.pushButton_3.clicked.connect(self.click_table3)
        self.ui.pushButton_4.clicked.connect(self.click_table4)

    def click_ontol(self):
        self.onto = OntoWindow()
        self.onto.show()

    def click_table1(self):
        self.table1 = TableWindow('Table1')
        self.table1.show()

    def click_table2(self):
        self.table2 = TableWindow('Table2')
        self.table2.show()

    def click_table3(self):
        self.table3 = TableWindow('Table3')
        self.table3.show()

    def click_table4(self):
        self.table4 = TableWindow('Table4')
        self.table4.show()

    def click_gisto(self):
        self.gisto = GistoWindow()
        self.gisto.show()


def main():
    app = QApplication([])
    application = MainWindow()
    application.show()

    sys.exit(app.exec())

if __name__ == '__main__':
    main()